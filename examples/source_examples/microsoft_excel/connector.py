# This is an example for how to work with the fivetran_connector_sdk module.
# It defines a method, which fetches the Microsoft Excel file from S3, processes it using pandas,openpyxl and python-calamine and upserts the data using the Connector SDK.
# See the Technical Reference documentation (https://fivetran.com/docs/connectors/connector-sdk/technical-reference#update)
# and the Best Practices documentation (https://fivetran.com/docs/connectors/connector-sdk/best-practices) for details

# Import required classes from fivetran_connector_sdk
from fivetran_connector_sdk import Connector
from fivetran_connector_sdk import Operations as op
from fivetran_connector_sdk import Logging as log

# Import required libraries
import boto3 # This is used to connect to S3
import pandas as pd # This is used to process the Excel file using pandas
from openpyxl import load_workbook # This is used to process the Excel file using openpyxl
import json
import os
import tempfile # This is used to create a temporary file


def create_s3_client(configuration: dict):
    """
    Create and return an S3 client using the provided configuration.
    Args:
        configuration (dict): A dictionary containing the AWS S3 credentials and file name.
    Returns:
        An S3 client object.
    """

    # Extract AWS credentials and region from the configuration
    access_key = configuration.get("aws_access_key_id")
    secret_key = configuration.get("aws_secret_access_key")
    region_name = configuration.get("region_name")

    # Return the S3 client using the credentials and region name from the configuration
    return boto3.client('s3',
                        aws_access_key_id=access_key,
                        aws_secret_access_key=secret_key,
                        region_name=region_name)


def download_excel_file(s3_client, bucket_name, file_key):
    """
    Download Excel file from S3 using a temporary file.
    Args:
        s3_client: The S3 client object.
        bucket_name (str): The name of the S3 bucket.
        file_key (str): The key of the file to download.
    Returns:
        str: The path to the temporary file where the Excel file is downloaded.
    """

    try:
        # Create a temporary file to store the downloaded Excel file
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_filename = temp_file.name

        # Download file from S3 to the temporary file
        s3_client.download_file(Bucket=bucket_name, Key=file_key, Filename=temp_filename)
        log.info(f"Downloaded file to temporary location: {temp_filename}")
        # Return the path to the temporary file
        return temp_filename
    except Exception as e:
        # Raise Runtime error if there is an error downloading the file
        raise RuntimeError(f"Error downloading the excel file: {e}")


def cleanup_temp_file(temp_filename):
    """
    Clean up the temporary file created during the download process.
    Args:
        temp_filename: The path to the temporary file.
    """

    # Remove the temporary file if it exists
    if os.path.exists(temp_filename):
        os.unlink(temp_filename)
        log.info(f"Temporary file {temp_filename} deleted.")


def upsert_using_pandas(temp_filename, state):
    """
    Process the Excel file using pandas and upsert the data into the destination table.
    You can modify the upsert logic and data manipulation here to suit your needs.
    This is suitable for small to medium-sized files as it loads the entire file into memory.
    This is not recommended for large files, because it can cause Out of Memory errors.
    Args:
        temp_filename: The path to the temporary file containing the Excel data.
        state:  a dictionary containing the state checkpointed during the prior sync.
    """

    # Read Excel file using pandas from the temporary file
    # This loads the entire Excel file into a pandas DataFrame
    # This is not recommended for large files, because it can cause Out of Memory errors
    df = pd.read_excel(temp_filename)

    # Iterate over the DataFrame rows and upsert each record
    for _, row in df.iterrows():
        # Convert the row to a dictionary
        # Each row is a dictionary with column names as keys
        record = row.to_dict()
        # The yield statement returns a generator object.
        # This generator will yield an upsert operation to the Fivetran connector.
        # The op.upsert method is called with two arguments:
        # - The first argument is the name of the table to upsert the data into, in this case, "excel_data".
        # - The second argument is a dictionary containing the data to be upserted.
        yield op.upsert(table="excel_data_pandas", data=record)

    # Save the progress by checkpointing the state. This is important for ensuring that the sync process can resume
    # from the correct position in case of next sync or interruptions.
    # Learn more about how and where to checkpoint by reading our best practices documentation
    # (https://fivetran.com/docs/connectors/connector-sdk/best-practices#largedatasetrecommendation).
    yield op.checkpoint(state)


def upsert_using_calamine(temp_filename, state):
    """
    Process the Excel file using python-calamine and upsert the data into the destination table.
    You can modify the upsert logic and data manipulation here to suit your needs.
    python-calamine is a Rust-backed library that loads the sheet into memory using efficient data structures.
    The pandas version 2.2 and above natively supports the calamine engine.
    This is recommended for large files because it has significantly lower memory overhead (7-9× faster than openpyxl)
    and better performance on low-resource systems.
    But, python-calamine also loads the entire file into the memory before iterating over the rows.
    This is why, python-calamine is not recommended for files above 1GB in size as it can cause memory overflow errors.
    Args:
        temp_filename: The path to the temporary file containing the Excel data.
        state:  a dictionary containing the state checkpointed during the prior sync.
    """

    # Read Excel file using pandas from the temporary file
    # This does not load the entire Excel file into the memory
    # This is highly recommended for large files, because it handles the data efficiently.
    # However, python-calamine is not recommended for files above 1GB in size as it loads entire file into the memory.
    df = pd.read_excel(temp_filename, engine="calamine")

    # Iterate over the DataFrame rows and upsert each record
    for _, row in df.iterrows():
        # Convert the row to a dictionary
        # Each row is a dictionary with column names as keys
        record = row.to_dict()
        # Upsert the record into the destination table
        yield op.upsert(table="excel_data_calamine", data=record)

    # Save the progress by checkpointing the state.
    yield op.checkpoint(state)


def upsert_using_openpyxl(temp_filename, state):
    """
    Process the Excel file using Openpyxl and upsert the data into the destination table.
    You can modify the upsert logic and data manipulation here to suit your needs.
    While openpyxl's default mode is memory-intensive (consuming up to 50× the file size in RAM),
    the read-only mode used here reduces memory by streaming rows, though at the cost of slower performance
    and limited features compared to python-calamine.
    openpyxl read-only mode does not load entire dataset into the memory but performs slower than python-calamine
    Args:
        temp_filename: The path to the temporary file containing the Excel data.
        state:  a dictionary containing the state checkpointed during the prior sync.
    """

    # Read Excel file using Openpyxl from the temporary file
    # This does not load the entire Excel file into the memory
    # This is recommended for large files, because it does not require loading the entire file into memory
    wb = load_workbook(filename=temp_filename, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    # Iterate over the rows and upsert each record
    for row in rows:
        # Convert the row to a dictionary
        # Each row is a dictionary with column names as keys
        record = dict(zip(headers, row))
        # Upsert the record into the destination table
        yield op.upsert(table="excel_data_openpyxl", data=record)

    # Save the progress by checkpointing the state.
    yield op.checkpoint(state)
    wb.close()


def schema(configuration: dict):
    """
    Define the schema function which lets you configure the schema your connector delivers.
    See the technical reference documentation for more details on the schema function:
    https://fivetran.com/docs/connectors/connector-sdk/technical-reference#schema
    Args:
        configuration: a dictionary that holds the configuration settings for the connector.
    """

    # Check if the required keys are present in the configuration
    required_keys = ["aws_access_key_id", "aws_secret_access_key", "region_name", "bucket_name", "file_name"]
    for key in required_keys:
        if key not in configuration:
            raise ValueError(f"Missing required configuration key: {key}")

    return [
        {
            "table": "excel_data_pandas", # Name of the table
            "primary_key": ["id"], # Primary key(s) of the table
            "columns": {
                "id": "INT",
                "name": "STRING",
                "email": "STRING",
                "age": "INT",
                "country": "STRING",
                "timestamp": "UTC_DATETIME",
            }
            # Columns not defined in schema will be inferred
        },
        {
            "table": "excel_data_calamine",  # Name of the table
            "primary_key": ["id"],  # Primary key(s) of the table
            "columns": {
                "id": "INT",
                "name": "STRING",
                "email": "STRING",
                "age": "INT",
                "country": "STRING",
                "timestamp": "UTC_DATETIME",
            }
            # Columns not defined in schema will be inferred
        },
        {
            "table": "excel_data_openpyxl",  # Name of the table
            "primary_key": ["id"],  # Primary key(s) of the table
            "columns": {
                "id": "INT",
                "name": "STRING",
                "email": "STRING",
                "age": "INT",
                "country": "STRING",
                "timestamp": "UTC_DATETIME",
            }
            # Columns not defined in schema will be inferred
        },
    ]


def update(configuration: dict, state: dict):
    """
    Define the update function, which is a required function, and is called by Fivetran during each sync.
    See the technical reference documentation for more details on the update function:
    https://fivetran.com/docs/connectors/connector-sdk/technical-reference#update
    Args:
        configuration: dictionary containing any secrets or payloads you configure when deploying the connector.
        state:  a dictionary containing the state checkpointed during the prior sync.
        The state dictionary is empty for the first sync or for any full re-sync.
    """
    log.warning("Example: Source Example - Microsoft Excel file")

    # Create S3 client
    s3_client = create_s3_client(configuration)

    # Get the bucket name and file key from the configuration
    bucket_name = configuration["bucket_name"]
    file_key = configuration.get("file_name")

    # Download the Excel file from S3 to a temporary file
    temp_filename = download_excel_file(s3_client, bucket_name, file_key)

    try:
        # There are three different ways to sync data from Excel file
        # Any of the three methods can be used to sync data from the Excel file
        # - python-calamine is recommended for the best balance of speed and memory efficiency
        #   However, python-calamine is not recommended for files above 1GB in size as it can cause memory overflow errors.
        # - openpyxl in read-only mode uses less memory but is significantly slower
        # - pandas with default engine is not recommended for large files as it loads everything into memory and can cause memory overflow errors

        # Process the Excel file using pandas and upsert the data
        yield from upsert_using_pandas(temp_filename, state)

        # Process the Excel file using python-calamine and upsert the data
        yield from upsert_using_calamine(temp_filename, state)

        # Process the Excel file using openpyxl and upsert the data
        yield from upsert_using_openpyxl(temp_filename, state)
    except Exception as e:
        # Raise Runtime error if there is an error processing the file
        raise RuntimeError(f"Error processing file {file_key}: {e}")
    finally:
        # Clean up the temporary file
        cleanup_temp_file(temp_filename)


# This creates the connector object that will use the update function defined in this connector.py file.
connector = Connector(update=update, schema=schema)

if __name__ == "__main__":
    # Check if the script is being run as the main module.
    # This is Python's standard entry method allowing your script to be run directly from the command line or IDE 'run' button.
    # This is useful for debugging while you write your code. Note this method is not called by Fivetran when executing your connector in production.
    # Please test using the Fivetran debug command prior to finalizing and deploying your connector.
    with open("configuration.json", 'r') as f:
        configuration = json.load(f)

    # Adding this code to your `connector.py` allows you to test your connector by running your file directly from your IDE:
    connector.debug(configuration=configuration)